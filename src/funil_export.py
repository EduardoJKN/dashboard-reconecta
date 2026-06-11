"""Exportação de relatórios do Funil da Reconecta (CSV, Excel, PDF)."""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from typing import Any, Literal

import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


MetricKind = Literal["money", "percent", "integer", "text"]

_XLS_MONEY = '"R$" #.##0,00'
# Formato Excel padrão (ponto decimal no code); exibe 73,91% no Excel pt-BR.
_XLS_PERCENT = "0.00%"
_XLS_INTEGER = "#.##0"
_XLS_DATE = "DD/MM/YYYY"

_WINE_RGB = "5C001E"
_HEADER_FILL = PatternFill("solid", fgColor=_WINE_RGB)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(size=10)
_TITLE_FONT = Font(bold=True, size=16, color=_WINE_RGB)
_SUBTITLE_FONT = Font(size=11, color="666666")
_SECTION_FONT = Font(bold=True, size=12, color=_WINE_RGB)
_KV_LABEL_FONT = Font(bold=True, size=10)
_THIN = Side(style="thin", color="CCCCCC")
_TABLE_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@dataclass
class FunilExportBundle:
    """Pacote de dados para exportação."""

    periodo_viz: str
    periodo_viz_label: str
    data_ini: date
    data_fim: date
    excluir_testes: bool
    atual: Any
    simulador: Any
    meta: Any
    calc_atual: dict[str, float]
    calc_sim: dict[str, float]
    calc_meta: dict[str, float]
    impactos: list[dict]
    periodos_cfg: dict


@dataclass(frozen=True)
class ComparativoRow:
    metrica: str
    kind: MetricKind
    atual: float
    simulador: float
    meta: float


# =============================================================================
# Formatação (exibição / exportação — não altera cálculos)
# =============================================================================

def fmt_brl(value: float | int | None) -> str:
    v = float(value or 0)
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_percent(value: float | int | None) -> str:
    """Taxa em fração (0,7391) → 73,91%."""
    v = float(value or 0)
    return f"{v * 100:.2f}%".replace(".", ",")


def fmt_int(value: float | int | None) -> str:
    return f"{int(float(value or 0)):,}".replace(",", ".")


def metric_kind(metric_name: str) -> MetricKind:
    name = metric_name.lower()
    if "%" in metric_name:
        return "percent"
    money_kw = (
        "r$", "investimento", "custo", "ticket", "montante",
        "receita", "ganho", "delta", "impacto", "diferença",
    )
    if any(k in name for k in money_kw):
        return "money"
    return "integer"


def fmt_value(value: float | int | None, metric_name: str) -> str:
    kind = metric_kind(metric_name)
    if kind == "money":
        return fmt_brl(value)
    if kind == "percent":
        return fmt_percent(value)
    if kind == "integer":
        return fmt_int(value)
    return str(value if value is not None else "")


def normalize_percent_for_excel(value: float | int | None) -> float:
    """Converte taxa para escala decimal do Excel (0,7391 → 73,91% com 0.00%).

    No Scenario as taxas são frações (0–1+). Valores > 10 assumem escala 0–100
    (ex.: 73,91 ou 235,29 → 0,7391 ou 2,3529).
    """
    if value is None:
        return 0.0
    v = float(value)
    if abs(v) > 10:
        return v / 100.0
    return v


def pdf_safe_text(text: str) -> str:
    """Seta unicode vira '->' no PDF; normaliza espaços."""
    t = str(text).replace("\u2192", "->").replace("→", "->").replace("−", "-")
    t = re.sub(r"\s+", " ", t).strip()
    return unicodedata.normalize("NFKC", t)


def _gargalo_top(impactos: list[dict]) -> dict | None:
    for item in impactos:
        if item.get("impacto", 0) > 0:
            return item
    return None


def _comparativo_rows(bundle: FunilExportBundle) -> list[ComparativoRow]:
    ca, cs, cm = bundle.calc_atual, bundle.calc_sim, bundle.calc_meta
    a, s, m = bundle.atual, bundle.simulador, bundle.meta
    return [
        ComparativoRow("Investimento", "money",
                       ca["investimento"], cs["investimento"], cm["investimento"]),
        ComparativoRow("Custo por Lead (R$)", "money",
                       a.custo_lead, s.custo_lead, m.custo_lead),
        ComparativoRow("Leads", "integer",
                       ca["leads"], cs["leads"], cm["leads"]),
        ComparativoRow("% Lead → Aplicação", "percent", a.pct_la, s.pct_la, m.pct_la),
        ComparativoRow("Aplicações", "integer",
                       ca["aplicacoes"], cs["aplicacoes"], cm["aplicacoes"]),
        ComparativoRow("% Aplicação → Agendamento", "percent",
                       a.pct_a_ag, s.pct_a_ag, m.pct_a_ag),
        ComparativoRow("Agendamentos", "integer",
                       ca["agendamentos"], cs["agendamentos"], cm["agendamentos"]),
        ComparativoRow("% Agendamento → Comparecimento", "percent",
                       a.pct_ag_c, s.pct_ag_c, m.pct_ag_c),
        ComparativoRow("Comparecimento", "integer",
                       ca["comparecimento"], cs["comparecimento"], cm["comparecimento"]),
        ComparativoRow("% Comparecimento → Venda", "percent",
                       a.pct_c_v, s.pct_c_v, m.pct_c_v),
        ComparativoRow("Vendas", "integer",
                       ca["vendas"], cs["vendas"], cm["vendas"]),
        ComparativoRow("Ticket Médio (R$)", "money", a.ticket, s.ticket, m.ticket),
        ComparativoRow("Montante", "money",
                       ca["montante"], cs["montante"], cm["montante"]),
        ComparativoRow("Receita", "money",
                       ca.get("receita", 0), cs.get("receita", 0), cm.get("receita", 0)),
    ]


# =============================================================================
# DataFrames
# =============================================================================

def build_export_resumo_df(bundle: FunilExportBundle) -> pd.DataFrame:
    gargalo = _gargalo_top(bundle.impactos)
    gargalo_txt = gargalo["label"] if gargalo else "Nenhum (funil alinhado à meta)"
    delta_sim = bundle.calc_sim["montante"] - bundle.calc_atual["montante"]
    delta_meta = bundle.calc_meta["montante"] - bundle.calc_atual["montante"]
    delta_rec_sim = (
        float(bundle.calc_sim.get("receita") or 0)
        - float(bundle.calc_atual.get("receita") or 0)
    )
    delta_rec_meta = (
        float(bundle.calc_meta.get("receita") or 0)
        - float(bundle.calc_atual.get("receita") or 0)
    )
    return pd.DataFrame(
        {
            "Campo": [
                "Relatório",
                "Período início",
                "Período fim",
                "Visualização",
                "Excluir testes nas aplicações",
                "Gargalo crítico",
                "Δ Montante Simulador - Atual",
                "Δ Montante Meta - Atual",
                "Δ Receita Simulador - Atual",
                "Δ Receita Meta - Atual",
            ],
            "Valor": [
                "Funil da Reconecta",
                bundle.data_ini.strftime("%d/%m/%Y"),
                bundle.data_fim.strftime("%d/%m/%Y"),
                bundle.periodo_viz_label,
                "Sim" if bundle.excluir_testes else "Não",
                gargalo_txt,
                fmt_brl(delta_sim),
                fmt_brl(delta_meta),
                fmt_brl(delta_rec_sim),
                fmt_brl(delta_rec_meta),
            ],
        }
    )


def build_export_comparativo_df(bundle: FunilExportBundle) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for row in _comparativo_rows(bundle):
        diff_sim = row.simulador - row.atual
        diff_meta = row.meta - row.atual
        records.append(
            {
                "Métrica": row.metrica,
                "Atual": row.atual,
                "Simulador": row.simulador,
                "Meta": row.meta,
                "Diferença Simulador vs Atual": diff_sim,
                "Diferença Meta vs Atual": diff_meta,
                "_kind": row.kind,
            }
        )
    return pd.DataFrame(records)


def build_export_gargalos_df(bundle: FunilExportBundle) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for idx, item in enumerate(bundle.impactos, start=1):
        label = item["label"]
        atual = float(item["atual"])
        meta = float(item["meta"])
        impacto = float(item.get("impacto") or 0)
        kind: MetricKind = "money" if item.get("is_money") else "percent"
        rows.append(
            {
                "Prioridade": idx,
                "Etapa": label,
                "Valor atual": atual,
                "Meta": meta,
                "Ganho potencial mensal": impacto,
                "_kind_atual": kind,
            }
        )
    if not rows:
        return pd.DataFrame(
            columns=[
                "Prioridade", "Etapa", "Valor atual", "Meta",
                "Ganho potencial mensal",
            ],
        )
    return pd.DataFrame(rows)


def build_export_metas_df(bundle: FunilExportBundle) -> pd.DataFrame:
    md = {
        "investimento": float(bundle.meta.investimento),
        "custo_lead": float(bundle.meta.custo_lead),
        "pct_la": float(bundle.meta.pct_la),
        "pct_a_ag": float(bundle.meta.pct_a_ag),
        "pct_ag_c": float(bundle.meta.pct_ag_c),
        "pct_c_v": float(bundle.meta.pct_c_v),
        "ticket": float(bundle.meta.ticket),
    }
    specs = [
        ("Investimento (mês)", md["investimento"], "money"),
        ("Custo por Lead (R$)", md["custo_lead"], "money"),
        ("% Lead → Aplicação", md["pct_la"], "percent"),
        ("% Aplicação → Agendamento", md["pct_a_ag"], "percent"),
        ("% Agendamento → Comparecimento", md["pct_ag_c"], "percent"),
        ("% Comparecimento → Venda", md["pct_c_v"], "percent"),
        ("Ticket Médio (R$)", md["ticket"], "money"),
    ]
    return pd.DataFrame(
        {
            "Métrica": [s[0] for s in specs],
            "Valor da meta": [s[1] for s in specs],
            "_kind": [s[2] for s in specs],
        }
    )


def build_export_dataframes(bundle: FunilExportBundle) -> dict[str, pd.DataFrame]:
    return {
        "Resumo Executivo": build_export_resumo_df(bundle),
        "Comparativo de Cenários": build_export_comparativo_df(bundle),
        "Gargalos e Prioridades": build_export_gargalos_df(bundle),
        "Metas Oficiais": build_export_metas_df(bundle),
    }


def build_export_dataframe(bundle: FunilExportBundle) -> pd.DataFrame:
    records = []
    for row in _comparativo_rows(bundle):
        diff_sim = row.simulador - row.atual
        diff_meta = row.meta - row.atual
        records.append(
            {
                "Métrica": row.metrica,
                "Atual": fmt_value(row.atual, row.metrica),
                "Simulador": fmt_value(row.simulador, row.metrica),
                "Meta": fmt_value(row.meta, row.metrica),
                "Diferença Simulador vs Atual": fmt_value(diff_sim, row.metrica),
                "Diferença Meta vs Atual": fmt_value(diff_meta, row.metrica),
            }
        )
    return pd.DataFrame(records)


# =============================================================================
# Excel — estilização
# =============================================================================

def _xls_num_format(kind: MetricKind) -> str | None:
    if kind == "money":
        return _XLS_MONEY
    if kind == "percent":
        return _XLS_PERCENT
    if kind == "integer":
        return _XLS_INTEGER
    return None


def _style_header_row(ws: Worksheet, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _TABLE_BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24


def _apply_body_style(ws: Worksheet, min_row: int, max_row: int, ncol: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _BODY_FONT
            cell.border = _TABLE_BORDER
            if c == 1:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autosize_columns(ws: Worksheet) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            if isinstance(val, date):
                max_len = max(max_len, 10)
            else:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 52)


def _set_num_cell(cell, value: float, kind: MetricKind) -> None:
    if kind == "percent":
        cell.value = normalize_percent_for_excel(value)
    else:
        cell.value = float(value or 0)
    nf = _xls_num_format(kind)
    if nf:
        cell.number_format = nf


def _merge_title_row(ws: Worksheet, row: int, text: str, font: Font, ncol: int = 5) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(vertical="center")


def _write_section_heading(ws: Worksheet, row: int, title: str, ncol: int = 5) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _SECTION_FONT
    return row + 1


def _write_kv_rows(
    ws: Worksheet,
    start_row: int,
    pairs: list[tuple[str, Any, MetricKind]],
) -> int:
    """Linhas label | valor com formatação por tipo."""
    r = start_row
    for label, value, kind in pairs:
        ws.cell(row=r, column=1, value=label).font = _KV_LABEL_FONT
        ws.cell(row=r, column=1).border = _TABLE_BORDER
        cell_v = ws.cell(row=r, column=2)
        if kind == "money":
            _set_num_cell(cell_v, float(value), "money")
        elif kind == "percent":
            _set_num_cell(cell_v, float(value), "percent")
        elif kind == "integer":
            _set_num_cell(cell_v, float(value), "integer")
        elif kind == "text":
            cell_v.value = value
        cell_v.border = _TABLE_BORDER
        cell_v.alignment = Alignment(
            horizontal="right" if kind != "text" else "left",
            vertical="center",
            wrap_text=True,
        )
        r += 1
    return r


def _write_table_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _TABLE_BORDER
    ws.row_dimensions[row].height = 22


def _write_resumo_executivo_sheet(ws: Worksheet, bundle: FunilExportBundle) -> None:
    """Capa executiva — espelha a estrutura do PDF."""
    gargalo = _gargalo_top(bundle.impactos)
    gargalo_txt = (
        gargalo["label"] if gargalo else "Nenhum (funil alinhado à meta)"
    )
    delta_sim = bundle.calc_sim["montante"] - bundle.calc_atual["montante"]
    delta_meta = bundle.calc_meta["montante"] - bundle.calc_atual["montante"]
    delta_rec_sim = (
        float(bundle.calc_sim.get("receita") or 0)
        - float(bundle.calc_atual.get("receita") or 0)
    )
    delta_rec_meta = (
        float(bundle.calc_meta.get("receita") or 0)
        - float(bundle.calc_atual.get("receita") or 0)
    )
    ncol = 5

    r = 1
    _merge_title_row(ws, r, "Funil da Reconecta", _TITLE_FONT, ncol)
    r += 1
    _merge_title_row(
        ws, r, "Relatório executivo de metas e gargalos", _SUBTITLE_FONT, ncol,
    )
    r += 2

    r = _write_section_heading(ws, r, "Contexto", ncol)
    r = _write_kv_rows(
        ws,
        r,
        [
            (
                "Período",
                f"{bundle.data_ini:%d/%m/%Y} a {bundle.data_fim:%d/%m/%Y}",
                "text",
            ),
            ("Visualização", bundle.periodo_viz_label, "text"),
            (
                "Excluir testes",
                "Sim" if bundle.excluir_testes else "Não",
                "text",
            ),
            ("Gargalo crítico", gargalo_txt, "text"),
        ],
    )
    r += 1

    r = _write_section_heading(ws, r, "Resumo", ncol)
    r = _write_kv_rows(
        ws,
        r,
        [
            ("Montante Atual", bundle.calc_atual["montante"], "money"),
            ("Montante Simulador", bundle.calc_sim["montante"], "money"),
            ("Montante Meta", bundle.calc_meta["montante"], "money"),
            ("Receita Atual", bundle.calc_atual.get("receita", 0), "money"),
            ("Receita Simulador", bundle.calc_sim.get("receita", 0), "money"),
            ("Receita Meta", bundle.calc_meta.get("receita", 0), "money"),
            ("Delta Montante Simulador - Atual", delta_sim, "money"),
            ("Delta Montante Meta - Atual", delta_meta, "money"),
            ("Delta Receita Simulador - Atual", delta_rec_sim, "money"),
            ("Delta Receita Meta - Atual", delta_rec_meta, "money"),
        ],
    )
    r += 1

    r = _write_section_heading(ws, r, "Principais oportunidades", ncol)
    opp_headers = ["Prioridade", "Etapa", "Ganho potencial mensal"]
    _write_table_header(ws, r, opp_headers)
    r += 1
    ranked = [x for x in bundle.impactos if x.get("impacto", 0) > 0][:8]
    if ranked:
        for pri, item in enumerate(ranked, start=1):
            ws.cell(row=r, column=1, value=pri)
            ws.cell(row=r, column=1).number_format = _XLS_INTEGER
            ws.cell(row=r, column=2, value=item["label"])
            _set_num_cell(
                ws.cell(row=r, column=3), float(item["impacto"]), "money",
            )
            _apply_body_style(ws, r, r, 3)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cell = ws.cell(row=r, column=1, value="Nenhuma oportunidade identificada.")
        cell.font = _BODY_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    r += 1

    r = _write_section_heading(ws, r, "Comparativo resumido", ncol)
    comp_headers = ["Métrica", "Atual", "Simulador", "Meta"]
    _write_table_header(ws, r, comp_headers)
    r += 1
    for row in _comparativo_rows(bundle):
        ws.cell(row=r, column=1, value=row.metrica)
        _set_num_cell(ws.cell(row=r, column=2), row.atual, row.kind)
        _set_num_cell(ws.cell(row=r, column=3), row.simulador, row.kind)
        _set_num_cell(ws.cell(row=r, column=4), row.meta, row.kind)
        _apply_body_style(ws, r, r, 4)
        r += 1

    _autosize_columns(ws)


def _write_comparativo_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    cols = [
        "Métrica", "Atual", "Simulador", "Meta",
        "Diferença Simulador vs Atual", "Diferença Meta vs Atual",
    ]
    ws.append(cols)
    for _, row in df.iterrows():
        kind: MetricKind = row["_kind"]
        ws.append([row["Métrica"]])
        r = ws.max_row
        for c, key in enumerate(
            ("Atual", "Simulador", "Meta",
             "Diferença Simulador vs Atual", "Diferença Meta vs Atual"),
            start=2,
        ):
            _set_num_cell(ws.cell(row=r, column=c), float(row[key]), kind)
    _style_header_row(ws, len(cols))
    _apply_body_style(ws, 2, ws.max_row, len(cols))
    _autosize_columns(ws)


def _write_gargalos_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    cols = ["Prioridade", "Etapa", "Valor atual", "Meta", "Ganho potencial mensal"]
    ws.append(cols)
    if df.empty:
        _style_header_row(ws, len(cols))
        return
    for _, row in df.iterrows():
        kind_a: MetricKind = row["_kind_atual"]
        ws.append([int(row["Prioridade"]), row["Etapa"]])
        r = ws.max_row
        ws.cell(row=r, column=1).number_format = _XLS_INTEGER
        _set_num_cell(ws.cell(row=r, column=3), float(row["Valor atual"]), kind_a)
        _set_num_cell(ws.cell(row=r, column=4), float(row["Meta"]), kind_a)
        _set_num_cell(
            ws.cell(row=r, column=5), float(row["Ganho potencial mensal"]), "money",
        )
    _style_header_row(ws, len(cols))
    _apply_body_style(ws, 2, ws.max_row, len(cols))
    _autosize_columns(ws)


def _write_metas_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    cols = ["Métrica", "Valor da meta"]
    ws.append(cols)
    for _, row in df.iterrows():
        kind: MetricKind = row["_kind"]
        ws.append([row["Métrica"]])
        r = ws.max_row
        _set_num_cell(ws.cell(row=r, column=2), float(row["Valor da meta"]), kind)
    _style_header_row(ws, len(cols))
    _apply_body_style(ws, 2, ws.max_row, len(cols))
    _autosize_columns(ws)


def export_funil_excel(bundle: FunilExportBundle) -> bytes:
    frames = build_export_dataframes(bundle)
    wb = Workbook()
    wb.remove(wb.active)

    ws_resumo = wb.create_sheet("Resumo Executivo")
    _write_resumo_executivo_sheet(ws_resumo, bundle)

    ws_comp = wb.create_sheet("Comparativo de Cenários")
    _write_comparativo_sheet(ws_comp, frames["Comparativo de Cenários"])

    ws_garg = wb.create_sheet("Gargalos e Prioridades")
    _write_gargalos_sheet(ws_garg, frames["Gargalos e Prioridades"])

    ws_metas = wb.create_sheet("Metas Oficiais")
    _write_metas_sheet(ws_metas, frames["Metas Oficiais"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# CSV
# =============================================================================

def export_funil_csv(bundle: FunilExportBundle) -> bytes:
    parts: list[str] = []
    garg = build_export_gargalos_df(bundle)
    if not garg.empty:
        garg_rows = []
        for _, r in garg.iterrows():
            kind: MetricKind = r["_kind_atual"]
            garg_rows.append(
                {
                    "Prioridade": int(r["Prioridade"]),
                    "Etapa": r["Etapa"],
                    "Valor atual": (
                        fmt_brl(r["Valor atual"]) if kind == "money"
                        else fmt_percent(r["Valor atual"])
                    ),
                    "Meta": (
                        fmt_brl(r["Meta"]) if kind == "money"
                        else fmt_percent(r["Meta"])
                    ),
                    "Ganho potencial mensal": fmt_brl(r["Ganho potencial mensal"]),
                }
            )
        garg_csv = pd.DataFrame(garg_rows)
    else:
        garg_csv = garg

    metas = build_export_metas_df(bundle)
    metas_csv = pd.DataFrame(
        {
            "Métrica": metas["Métrica"],
            "Valor da meta": [
                fmt_value(v, m) for v, m in zip(metas["Valor da meta"], metas["Métrica"])
            ],
        }
    )

    for title, df in (
        ("Resumo Executivo", build_export_resumo_df(bundle)),
        ("Comparativo de Cenários", build_export_dataframe(bundle)),
        ("Gargalos e Prioridades", garg_csv),
        ("Metas Oficiais", metas_csv),
    ):
        parts.append(f"### {title}")
        parts.append(df.to_csv(index=False, sep=";"))
        parts.append("")
    return ("\ufeff" + "\n".join(parts)).encode("utf-8")


# =============================================================================
# PDF — relatório executivo
# =============================================================================

class _FunilPDF(FPDF):
    def footer(self) -> None:
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 8, pdf_safe_text(f"Página {self.page_no()}"), align="C")
        self.set_text_color(0, 0, 0)


def _pdf_section_title(pdf: FPDF, title: str) -> None:
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(92, 0, 30)
    pdf.cell(0, 8, pdf_safe_text(title), ln=True)
    pdf.set_text_color(0, 0, 0)


def _pdf_kv_block(pdf: FPDF, pairs: list[tuple[str, str]]) -> None:
    col_w = (pdf.w - pdf.l_margin - pdf.r_margin) / 2
    for label, value in pairs:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(col_w, 6, pdf_safe_text(label), ln=0)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(col_w, 6, pdf_safe_text(value), ln=True)


def _pdf_table(
    pdf: FPDF,
    headers: list[str],
    rows: list[list[str]],
    col_widths: list[float],
) -> None:
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(92, 0, 30)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        align = "C" if h == "Prioridade" else "C"
        pdf.cell(col_widths[i], 7, pdf_safe_text(h), border=1, fill=True, align=align)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for row in rows:
        if pdf.get_y() > pdf.h - 22:
            pdf.add_page()
        for i, cell in enumerate(row):
            txt = pdf_safe_text(str(cell))
            align = "L"
            if headers[i] == "Prioridade":
                align = "C"
            elif i > 0 and headers[i] != "Etapa" and headers[i] != "Métrica":
                align = "R"
            pdf.cell(col_widths[i], 6, txt, border=1, align=align)
        pdf.ln()


def export_funil_pdf(bundle: FunilExportBundle) -> bytes:
    gargalo = _gargalo_top(bundle.impactos)
    delta_sim = bundle.calc_sim["montante"] - bundle.calc_atual["montante"]
    delta_meta = bundle.calc_meta["montante"] - bundle.calc_atual["montante"]
    delta_rec_sim = (
        float(bundle.calc_sim.get("receita") or 0)
        - float(bundle.calc_atual.get("receita") or 0)
    )
    delta_rec_meta = (
        float(bundle.calc_meta.get("receita") or 0)
        - float(bundle.calc_atual.get("receita") or 0)
    )

    pdf = _FunilPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()

    usable = pdf.w - pdf.l_margin - pdf.r_margin

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(92, 0, 30)
    pdf.cell(0, 10, pdf_safe_text("Funil da Reconecta"), ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, pdf_safe_text("Relatório executivo de metas e gargalos"), ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    _pdf_section_title(pdf, "Contexto")
    _pdf_kv_block(
        pdf,
        [
            ("Período:", f"{bundle.data_ini:%d/%m/%Y} a {bundle.data_fim:%d/%m/%Y}"),
            ("Visualização:", bundle.periodo_viz_label),
            ("Excluir testes:", "Sim" if bundle.excluir_testes else "Não"),
        ],
    )

    _pdf_section_title(pdf, "Resumo")
    gargalo_txt = (
        f"{gargalo['label']} (+ {fmt_brl(gargalo['impacto'])}/mês)"
        if gargalo else "Nenhum (alinhado à meta)"
    )
    _pdf_kv_block(
        pdf,
        [
            ("Montante Atual:", fmt_brl(bundle.calc_atual["montante"])),
            ("Montante Simulador:", fmt_brl(bundle.calc_sim["montante"])),
            ("Montante Meta:", fmt_brl(bundle.calc_meta["montante"])),
            ("Receita Atual:", fmt_brl(bundle.calc_atual.get("receita", 0))),
            ("Receita Simulador:", fmt_brl(bundle.calc_sim.get("receita", 0))),
            ("Receita Meta:", fmt_brl(bundle.calc_meta.get("receita", 0))),
            ("Delta Montante Sim. - Atual:", fmt_brl(delta_sim)),
            ("Delta Montante Meta - Atual:", fmt_brl(delta_meta)),
            ("Delta Receita Sim. - Atual:", fmt_brl(delta_rec_sim)),
            ("Delta Receita Meta - Atual:", fmt_brl(delta_rec_meta)),
            ("Gargalo crítico:", gargalo_txt),
        ],
    )

    ranked = [x for x in bundle.impactos if x.get("impacto", 0) > 0][:8]
    _pdf_section_title(pdf, "Oportunidades")
    if ranked:
        opp_rows = [
            [str(i), item["label"], fmt_brl(item["impacto"])]
            for i, item in enumerate(ranked, start=1)
        ]
        _pdf_table(
            pdf,
            ["Prioridade", "Etapa", "Ganho potencial mensal"],
            opp_rows,
            [16, usable - 16 - 44, 44],
        )
    else:
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, pdf_safe_text("Nenhuma oportunidade identificada."), ln=True)

    _pdf_section_title(pdf, "Comparativo de cenários")
    comp_rows = [
        [
            row.metrica,
            fmt_value(row.atual, row.metrica),
            fmt_value(row.simulador, row.metrica),
            fmt_value(row.meta, row.metrica),
        ]
        for row in _comparativo_rows(bundle)
    ]
    w_metrica = usable * 0.38
    w_val = (usable - w_metrica) / 3
    _pdf_table(
        pdf,
        ["Métrica", "Atual", "Simulador", "Meta"],
        comp_rows,
        [w_metrica, w_val, w_val, w_val],
    )

    out = pdf.output()
    if isinstance(out, (bytes, bytearray)):
        return bytes(out)
    return out.encode("latin-1")
